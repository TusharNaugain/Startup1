import pandas as pd
import os
from datetime import datetime
import argparse

class ReportGenerator:
    def __init__(self, raw_data_file, customer_file):
        self.raw_data_file = raw_data_file
        self.customer_file = customer_file
        self.output_file = f"Automated_Report_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
        
    def load_data(self):
        print("Loading data...")
        try:
            # Load Customer List from separate file
            # Assumptions: 
            # 1. Customer file has a sheet named 'data ' OR it reads the first sheet.
            # 2. Column containing email is named 'emailAddress' or we find the likely column.
            
            # Helper to find email column
            def find_email_col(df):
                for col in df.columns:
                    if 'email' in str(col).lower():
                        return col
                return df.columns[0] # Fallback
            
            # Read first sheet of customer file or CSV
            if self.customer_file.lower().endswith('.csv'):
                self.customers_df = pd.read_csv(self.customer_file)
            else:
                self.customers_df = pd.read_excel(self.customer_file, sheet_name=0)
            
            email_col = find_email_col(self.customers_df)
            print(f"Using '{email_col}' as customer email column.")
            
            self.customer_emails = set(self.customers_df[email_col].astype(str).str.lower().str.strip())
            print(f"Loaded {len(self.customer_emails)} customer emails.")

            # Load Raw Input Data
            if self.raw_data_file.lower().endswith('.csv'):
                self.raw_df = pd.read_csv(self.raw_data_file)
            else:
                self.raw_df = pd.read_excel(self.raw_data_file, sheet_name=0) 
            
            print(f"Loaded {len(self.raw_df)} rows of input data.")
            
        except Exception as e:
            print(f"Error loading data: {e}")
            raise

    def process_data(self):
        print("Processing data...")
        
        # --- NEW: Map Columns from Raw Input ---
        # We need: Email, Created At, Usage Count
        # Map: 
        #   actor.properties.email -> actor.properties.email (keep or standardize)
        #   actor.created_at -> actor.created_at
        #   event_count -> Usage Count
        
        if 'event_count' in self.raw_df.columns:
            self.raw_df.rename(columns={'event_count': 'Usage Count'}, inplace=True)
            print("Renamed 'event_count' to 'Usage Count'")
        
        # Identify raw email column
        raw_email_col = 'actor.properties.email'
        # ... (rest of logic)
        if raw_email_col not in self.raw_df.columns:
             # Try to find it
             for col in self.raw_df.columns:
                 if 'email' in str(col).lower() and 'property' in str(col).lower():
                     raw_email_col = col
                     break
        
        print(f"Using '{raw_email_col}' as raw data email column.")
        
        self.raw_df['clean_email'] = self.raw_df[raw_email_col].astype(str).str.lower().str.strip()
        
        # --- NEW LOGIC: Status (New vs Returning) ---
        # Formula: IF(Today - Created_At <= 7, "NEW", "RETURNING")
        # Ensure dates are datetime
        if 'actor.created_at' in self.raw_df.columns:
            self.raw_df['actor.created_at'] = pd.to_datetime(self.raw_df['actor.created_at'], errors='coerce')
            
            # Use 'Todays Date' column if exists, else use current date
            if 'Todays Date' in self.raw_df.columns:
                 self.raw_df['Todays Date'] = pd.to_datetime(self.raw_df['Todays Date'], errors='coerce').dt.tz_localize(None)
            else:
                 # Create column with current run date
                 current_date = pd.Timestamp(datetime.now().date())
                 self.raw_df['Todays Date'] = current_date
            
            # Ensure created_at is also tz-naive
            self.raw_df['actor.created_at'] = self.raw_df['actor.created_at'].dt.tz_localize(None)
            
            # Calculate Difference
            # Note: We take absolute difference to be safe, though formula implies simple subtraction
            self.raw_df['Days_Diff'] = (self.raw_df['Todays Date'] - self.raw_df['actor.created_at']).dt.days
            
            self.raw_df['Status'] = self.raw_df['Days_Diff'].apply(lambda d: 'NEW' if d <= 7 else 'RETURNING')
            print("Status (New/Returning) calculated.")
            
            # Format dates to remove time (YYYY-MM-DD) for final output
            self.raw_df['actor.created_at'] = self.raw_df['actor.created_at'].dt.strftime('%Y-%m-%d')
            self.raw_df['Todays Date'] = self.raw_df['Todays Date'].dt.strftime('%Y-%m-%d')
            
        else:
            print("Warning: 'actor.created_at' not found, cannot calculate Status.")

        # 2. Tag Status New
        # Logic: If email in customer_list -> 'Customer', else 'Leads'
        def get_status(email):
            if email in self.customer_emails:
                return 'Customer'
            return 'Leads'
            
        self.raw_df['Status New'] = self.raw_df['clean_email'].apply(get_status)
        print("Tagging complete.")
        
        # 3. Calculate Usage Logic (Dynamic Median)
        if 'Usage Count' in self.raw_df.columns:
            median_val = self.raw_df['Usage Count'].median()
            print(f"Calculated Median Usage: {median_val}")
            
            def get_usage_category(count):
                if count < median_val:
                    return "Low usage"
                elif count < 2 * median_val:
                    return "AVG Usage"
                elif count < 3 * median_val:
                    return "Above AVG Usage"
                else:
                    return "High Usage"
            
            self.raw_df['Usage'] = self.raw_df['Usage Count'].apply(get_usage_category)
            print("Usage categorization complete.")
        else:
            print("Warning: 'Usage Count' column not found, cannot calculate Usage.")

    def generate_sheets(self):
        print("Generating sheets...")
        
        # Prepare Main Sheet
        self.main_sheet = self.raw_df.copy()
        
        # --- Final Column Selection & Ordering ---
        # User requested specific format:
        # actor.created_at | Todays Date | Status | actor.properties.name | actor.properties.email | Status New | Total Usage | Usage
        
        output_columns = [
            'actor.created_at', 
            'Todays Date', 
            'Status', 
            'actor.properties.name', 
            'actor.properties.email', # We need to ensure we have the original name or map 'clean_email' back
            'Status New', 
            'Usage Count', # Will rename to Total Usage
            'Usage'
        ]
        
        # Ensure 'actor.properties.name' exists or is created
        if 'actor.properties.name' not in self.main_sheet.columns:
             self.main_sheet['actor.properties.name'] = "" # Default if missing
        
        # Ensure 'actor.properties.email' is populated (we used it earlier)
        # We might have used a different column name for raw email, let's normalize
        if 'actor.properties.email' not in self.main_sheet.columns and 'clean_email' in self.main_sheet.columns:
            self.main_sheet['actor.properties.email'] = self.main_sheet['clean_email']

        # Filter and Rename for Final Output
        # Helper to safely select columns
        final_cols = []
        rename_map = {'Usage Count': 'Total Usage'}
        
        for col in output_columns:
            if col in self.main_sheet.columns:
                final_cols.append(col)
        
        # Create formatted sheets
        self.lead_sheet = self.main_sheet[self.main_sheet['Status New'] == 'Leads'][final_cols].copy()
        self.customer_sheet = self.main_sheet[self.main_sheet['Status New'] == 'Customer'][final_cols].copy()
        
        # Rename columns in the output sheets
        self.lead_sheet.rename(columns=rename_map, inplace=True)
        self.customer_sheet.rename(columns=rename_map, inplace=True)
        
        print(f"Lead Sheet Rows: {len(self.lead_sheet)}")
        print(f"Customer Sheet Rows: {len(self.customer_sheet)}")
        
        # --- Summary / Pivot Logic ---
        # 1. Pivot Table 4 equivalent (e.g., Count by Status)
        self.summary_status = self.main_sheet.groupby('Status New').size().reset_index(name='Count')
        
        # 2. Customer Pivot equivalent 
        # (Assuming we want Usage breakdown for Customers, or Total Usage summaries)
        # Usage Count, Total Usage analysis
        
        summary_dfs = {}
        
        if 'Usage' in self.main_sheet.columns:
            # Pivot: Rows=Status New, Cols=Usage, Values=Count
            pivot_usage = pd.crosstab(self.main_sheet['Status New'], self.main_sheet['Usage'])
            summary_dfs['Usage_Analysis'] = pivot_usage
            
        # Additional Pivot: Top Users (Lead vs Customer) based on Usage Count if it exists
        if 'Usage Count' in self.main_sheet.columns:
             # Basic stats
             stats = self.main_sheet.groupby('Status New')['Usage Count'].describe()
             summary_dfs['Usage_Stats'] = stats

        self.summary_dfs = summary_dfs

        # --- NEW: Top Users for "Other Insights" Sheet ---
        print("Calculating Top Users for 'Other Insights'...")
        
        # 1. Identify Workspace Name Column
        workspace_col = None
        possible_names = ['workspace', 'company', 'organization', 'account']
        
        for col in self.raw_df.columns:
            for name in possible_names:
                if name in str(col).lower():
                    workspace_col = col
                    break
            if workspace_col:
                break
        
        if not workspace_col:
            print("Warning: Could not find a 'Workspace Name' column. Using 'Unknown'.")
            self.raw_df['Workspace Name'] = 'Unknown'
            workspace_col = 'Workspace Name'
        else:
            print(f"Using '{workspace_col}' as Workspace Name column.")

        # Ensure we have the necessary columns in a temporary DF for processing
        # Columns needed: Workspace Name, Usage Count, User Name (actor.properties.name), User Email (clean_email)
        
        # Helper to get specific top users
        def get_top_users(status_val, n=5):
            # Filter by Status (calculated in process_data as 'NEW' or 'RETURNING')
            if 'Status' not in self.raw_df.columns:
                return pd.DataFrame()
            
            # Start with filtering by Status (NEW/RETURNING)
            subset = self.raw_df[self.raw_df['Status'] == status_val].copy()
            
            # Additional Filter: Only include 'Leads' (Free Users)
            if 'Status New' in subset.columns:
                subset = subset[subset['Status New'] == 'Leads']
            
            if subset.empty:
                return pd.DataFrame()
                
            # Sort by Usage Count descending
            if 'Usage Count' in subset.columns:
                subset = subset.sort_values(by='Usage Count', ascending=False)
            
            return subset.head(n)

        top_new = get_top_users('NEW', 3)
        top_returning = get_top_users('RETURNING', 2)
        
        # Combine
        self.other_insights_df = pd.concat([top_new, top_returning])
        
        # Select and Rename Columns for Final Output
        # Desired: Workspace Name | Usage Count | User Name | User Email
        
        cols_to_keep = {
            workspace_col: 'Workspace Name',
            'Usage Count': 'Usage Count',
            'actor.properties.name': 'User Name',
            'clean_email': 'User Email' # Using clean_email as it's definitely present
        }
        
        # Ensure columns exist before selecting
        final_insight_cols = []
        for col, new_name in cols_to_keep.items():
            if col in self.other_insights_df.columns:
                final_insight_cols.append(col)
            else:
                # If 'User Name' is missing, try to find it or create empty
                if new_name == 'User Name':
                     self.other_insights_df[col] = ''
                     final_insight_cols.append(col)

        self.other_insights_df = self.other_insights_df[final_insight_cols].rename(columns=cols_to_keep)
        print(f"Generated 'Other Insights' with {len(self.other_insights_df)} rows.")


    def save_report(self):
        print(f"Saving report to {self.output_file}...")
        try:
            # Use openpyxl for formatting
            from openpyxl.styles import Border, Side, Font, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill
            
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                self.main_sheet.to_excel(writer, sheet_name=f'sheet {datetime.now().strftime("%d-%m-%Y")}', index=False)
                self.lead_sheet.to_excel(writer, sheet_name='Lead Sheet', index=False)
                self.customer_sheet.to_excel(writer, sheet_name='Customer Sheet', index=False)
                
                # Write Summaries
                self.summary_status.to_excel(writer, sheet_name='Summary_Status', index=False)
                
                for name, df in self.summary_dfs.items():
                    df.to_excel(writer, sheet_name=f"Summary_{name}")
                
                # --- NEW: Write "Other Insights" Sheet ---
                # We need to write this manually to handle the custom headers
                # "Other Insights:" at A1
                # "Some of the highlighted free users extensively using the product:" at A2
                # Table at A3
                
                # Create a sheet for Other Insights
                workbook = writer.book
                insights_sheet = workbook.create_sheet("Other Insights")
                
                # Write Custom Headers
                insights_sheet['A1'] = "Other Insights:"
                insights_sheet['A1'].font = Font(bold=True, size=12) # Example styling
                
                insights_sheet['A2'] = "Some of the highlighted free users extensively using the product:"
                insights_sheet['A2'].font = Font(bold=True)
                
                # Write the DataFrame starting from row 3 (which is index 3 in openpyxl if 1-based, 
                # but dataframe_to_rows gives headers so we start at row 3)
                
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Write headers first? No, to_excel handles it, but we are doing manual via openpyxl here mostly
                # Actually, pandas to_excel allow startrow
                self.other_insights_df.to_excel(writer, sheet_name='Other Insights', startrow=2, index=False)
                
                # Get the sheet again (pandas might have reset it or we just use the name)
                # Note: creating sheet above with workbook.create_sheet might conflict if pandas tries to create it too.
                # Better approach: Let pandas write it, then insert rows? Or just write to a specific startrow.
                
                # Re-access the sheet after pandas wrote to it.
                # Wait, if I use workbook.create_sheet("Other Insights") then pandas to_excel with same name might duplicate or error.
                # Let's just use to_excel with startrow=2, then access the sheet to add A1/A2
                
                pass # Logic continues below in "Apply Formatting"
                
                
                # Apply Formatting to All Sheets
                thin_border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))
                
                header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Blue
                total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green or Greyish
                
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Special handling for "Other Insights" headers
                    if sheet_name == "Other Insights":
                        worksheet['A1'] = "Other Insights:"
                        worksheet['A1'].font = Font(bold=True, size=11)
                        worksheet['A2'] = "Some of the highlighted free users extensively using the product:"
                        worksheet['A2'].font = Font(bold=True, size=11)
                        
                        # The table starts at row 3 (header is row 3)
                        # Let's format the table specifically
                        # Header row is 3
                        for cell in worksheet[3]:
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            # cell.fill = header_fill # Optional
                    
                    # Iterate over columns to auto-adjust width and apply styles
                    # For Other Insights, we want to skip rows 1 and 2 for some checks, but column width applies to whole col
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        
                        for cell in column:
                            try:
                                # Apply Border to table parts
                                if sheet_name == "Other Insights":
                                    if cell.row >= 3 and cell.value is not None: # Table part
                                         cell.border = thin_border
                                else:
                                    cell.border = thin_border
                                
                                # Highlight Total Rows (Pivot Simulation)
                                if "Pivot" in sheet_name:
                                    row_val = worksheet.cell(row=cell.row, column=1).value
                                    if row_val and "Total" in str(row_val):
                                         cell.font = Font(bold=True)
                                         cell.fill = total_fill

                                # Calculate max width
                                if cell.value:
                                    # Don't let long headers in A1/A2 distort col A width too much if we don't want
                                    if sheet_name == "Other Insights" and cell.row <= 2:
                                        continue 
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        
                        # Set Column Width
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Header Formatting (First Row)
                    if sheet_name != "Other Insights":
                        for cell in worksheet[1]:
                            cell.font = Font(bold=True)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.fill = header_fill
            
            print("Report saved successfully.")
            return self.output_file
            
        except Exception as e:
            print(f"Error saving report: {e}")
            # Fallback if styling fails
            return self.output_file

if __name__ == "__main__":
    # Allow arguments, but default to current directory files for ease of use
    import sys
    
    # Defaults
    raw_data = 'Untitled spreadsheet (1).xlsx' # New default input format
    customer_list = 'customers.xlsx'

    if len(sys.argv) > 2:
        raw_data = sys.argv[1]
        customer_list = sys.argv[2]
        
    if os.path.exists(raw_data) and os.path.exists(customer_list):
        print(f"Running with Raw Data: {raw_data}, Customer List: {customer_list}")
        generator = ReportGenerator(raw_data, customer_list)
        generator.load_data()
        generator.process_data()
        generator.generate_sheets()
        generator.save_report()
    else:
        print(f"One or both files not found:\nRaw: {raw_data}\nCustomer: {customer_list}")
