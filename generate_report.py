import pandas as pd
import os
from datetime import datetime
import argparse
import smtplib
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage


class ReportGenerator:
    def __init__(self, raw_data_file, customer_file):
        self.raw_data_file = raw_data_file
        self.customer_file = customer_file
        self.output_file = f"Automated_Report_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

    # ------------------------------------------------------------------
    # STEP 1: Load Data
    # ------------------------------------------------------------------
    def load_data(self):
        print("Loading data...")
        try:
            def find_email_col(df):
                for col in df.columns:
                    if 'email' in str(col).lower():
                        return col
                return df.columns[0]

            if self.customer_file.lower().endswith('.csv'):
                self.customers_df = pd.read_csv(self.customer_file)
            else:
                self.customers_df = pd.read_excel(self.customer_file, sheet_name=0)

            email_col = find_email_col(self.customers_df)
            print(f"Using '{email_col}' as customer email column.")
            self.customer_emails = set(
                self.customers_df[email_col].astype(str).str.lower().str.strip()
            )
            print(f"Loaded {len(self.customer_emails)} customer emails.")

            if self.raw_data_file.lower().endswith('.csv'):
                self.raw_df = pd.read_csv(self.raw_data_file)
            else:
                self.raw_df = pd.read_excel(self.raw_data_file, sheet_name=0)

            print(f"Loaded {len(self.raw_df)} rows of input data.")

        except Exception as e:
            print(f"Error loading data: {e}")
            raise

    # ------------------------------------------------------------------
    # STEP 2: Process / Enrich Data
    # ------------------------------------------------------------------
    def process_data(self):
        print("Processing data...")

        # Rename event_count → Usage Count
        if 'event_count' in self.raw_df.columns:
            self.raw_df.rename(columns={'event_count': 'Usage Count'}, inplace=True)
            print("Renamed 'event_count' to 'Usage Count'")

        # Identify the email column in raw data
        raw_email_col = 'actor.properties.email'
        if raw_email_col not in self.raw_df.columns:
            for col in self.raw_df.columns:
                if 'email' in str(col).lower() and 'property' in str(col).lower():
                    raw_email_col = col
                    break
        print(f"Using '{raw_email_col}' as raw data email column.")

        self.raw_df['clean_email'] = (
            self.raw_df[raw_email_col].astype(str).str.lower().str.strip()
        )

        # NEW / RETURNING based on created_at date
        if 'actor.created_at' in self.raw_df.columns:
            self.raw_df['actor.created_at'] = pd.to_datetime(
                self.raw_df['actor.created_at'], errors='coerce'
            )

            if 'Todays Date' in self.raw_df.columns:
                self.raw_df['Todays Date'] = (
                    pd.to_datetime(self.raw_df['Todays Date'], errors='coerce')
                    .dt.tz_localize(None)
                )
            else:
                self.raw_df['Todays Date'] = pd.Timestamp(datetime.now().date())

            self.raw_df['actor.created_at'] = self.raw_df['actor.created_at'].dt.tz_localize(None)
            self.raw_df['Days_Diff'] = (
                self.raw_df['Todays Date'] - self.raw_df['actor.created_at']
            ).dt.days
            self.raw_df['Status'] = self.raw_df['Days_Diff'].apply(
                lambda d: 'NEW' if d <= 7 else 'RETURNING'
            )
            print("Status (New/Returning) calculated.")

            self.raw_df['actor.created_at'] = self.raw_df['actor.created_at'].dt.strftime('%Y-%m-%d')
            self.raw_df['Todays Date'] = self.raw_df['Todays Date'].dt.strftime('%Y-%m-%d')
        else:
            print("Warning: 'actor.created_at' not found, cannot calculate Status.")

        # Leads vs Customer tag
        def get_status_new(email):
            return 'Customer' if email in self.customer_emails else 'Leads'

        self.raw_df['Status New'] = self.raw_df['clean_email'].apply(get_status_new)
        print("Leads/Customer tagging complete.")

        # Usage buckets (dynamic median)
        if 'Usage Count' in self.raw_df.columns:
            median_val = self.raw_df['Usage Count'].median()
            print(f"Calculated Median Usage: {median_val}")

            def get_usage_category(count):
                if count < median_val:
                    return "Low usage"
                elif count < 2 * median_val:
                    return "AVG Usage"
                elif count < 3 * median_val:
                    return "Above AVG Usage"
                else:
                    return "High Usage"

            self.raw_df['Usage'] = self.raw_df['Usage Count'].apply(get_usage_category)
            print("Usage categorization complete.")
        else:
            print("Warning: 'Usage Count' column not found, cannot calculate Usage.")

    # ------------------------------------------------------------------
    # STEP 3: Generate Sheets (Lead, Customer + their Pivots)
    # ------------------------------------------------------------------
    def generate_sheets(self):
        print("Generating sheets...")
        self.main_sheet = self.raw_df.copy()

        output_columns = [
            'actor.created_at',
            'Todays Date',
            'Status',
            'actor.properties.name',
            'actor.properties.email',
            'Status New',
            'Usage Count',
            'Usage'
        ]

        if 'actor.properties.name' not in self.main_sheet.columns:
            self.main_sheet['actor.properties.name'] = ""

        if 'actor.properties.email' not in self.main_sheet.columns:
            self.main_sheet['actor.properties.email'] = self.main_sheet['clean_email']

        final_cols = [c for c in output_columns if c in self.main_sheet.columns]
        rename_map = {'Usage Count': 'Total Usage'}

        # Split into Lead / Customer
        self.lead_sheet = (
            self.main_sheet[self.main_sheet['Status New'] == 'Leads'][final_cols]
            .copy()
            .rename(columns=rename_map)
        )
        self.customer_sheet = (
            self.main_sheet[self.main_sheet['Status New'] == 'Customer'][final_cols]
            .copy()
            .rename(columns=rename_map)
        )

        print(f"Lead Sheet Rows: {len(self.lead_sheet)}")
        print(f"Customer Sheet Rows: {len(self.customer_sheet)}")

        # Build pivot tables for Lead and Customer sheets
        self.lead_pivot_rows = self._build_pivot_rows(self.lead_sheet)
        self.customer_pivot_rows = self._build_pivot_rows(self.customer_sheet)

        print(f"Lead Pivot Rows: {len(self.lead_pivot_rows)}")
        print(f"Customer Pivot Rows: {len(self.customer_pivot_rows)}")

        # Other Insights (top free users)
        self._build_other_insights()

    # ------------------------------------------------------------------
    # Pivot Builder
    # Produces a flat list of rows that match the screenshot layout:
    #   Status  |  Usage                 |  actor.properties.email  |  SUM of Total Usage
    #   NEW      |  Above AVG Usage       |  user@example.com        |  10
    #   NEW      |  Above AVG Usage       |  user2@example.com       |  11
    #            |  Above AVG Usage Total |                          |  21
    #   NEW      |  AVG Usage             |  ...                     |  ...
    #            |  AVG Usage Total       |                          |  18
    #   NEW Total|                        |                          |  46
    #   RETURNING|  ...                   |  ...                     |  ...
    #   Grand Total                                                   |  59
    # ------------------------------------------------------------------
    def _build_pivot_rows(self, df: pd.DataFrame) -> list:
        """
        Returns a list of dicts representing each row of the pivot table.
        Columns: Status | Usage | actor.properties.email | SUM of Total Usage
        """
        if 'Status' not in df.columns or 'Usage' not in df.columns or 'Total Usage' not in df.columns:
            print("Warning: Required columns for pivot missing. Skipping pivot build.")
            return []

        email_col = 'actor.properties.email'
        if email_col not in df.columns:
            email_col = 'clean_email' if 'clean_email' in df.columns else None

        rows = []
        grand_total = 0

        # Usage display order (matches screenshot top-to-bottom order within each Status)
        USAGE_ORDER = ["Above AVG Usage", "High Usage", "AVG Usage", "Low usage"]

        for status_val in sorted(df['Status'].dropna().unique()):
            status_df = df[df['Status'] == status_val]
            status_total = 0
            first_status_row = True  # Only show Status label on first row of that group

            # Get usage categories present, preserving preferred order
            usages_present = [u for u in USAGE_ORDER if u in status_df['Usage'].values]
            # Append any other categories not in USAGE_ORDER
            other_usages = [u for u in status_df['Usage'].dropna().unique() if u not in USAGE_ORDER]
            usages_present += sorted(other_usages)

            for usage_val in usages_present:
                usage_df = status_df[status_df['Usage'] == usage_val].copy()
                usage_total = 0
                first_usage_row = True  # Show Usage label on first row of that usage group

                if email_col:
                    # Aggregate: group by email (same email can appear multiple times)
                    email_summary = (
                        usage_df.groupby(email_col)['Total Usage']
                        .sum()
                        .reset_index()
                        .sort_values('Total Usage', ascending=False)
                    )

                    for _, row in email_summary.iterrows():
                        email = row[email_col]
                        total = row['Total Usage']
                        usage_total += total

                        rows.append({
                            'Status':                   status_val if first_status_row else '',
                            'Usage':                    usage_val if first_usage_row else '',
                            'actor.properties.email':   email,
                            'SUM of Total Usage':       total,
                            '_row_type':                'data'
                        })
                        first_status_row = False
                        first_usage_row = False
                else:
                    # No email column — just one aggregate row per usage bucket
                    total = usage_df['Total Usage'].sum()
                    usage_total = total
                    rows.append({
                        'Status':               status_val if first_status_row else '',
                        'Usage':                usage_val if first_usage_row else '',
                        'actor.properties.email': '',
                        'SUM of Total Usage':   total,
                        '_row_type':            'data'
                    })
                    first_status_row = False

                # Usage Sub-Total row
                rows.append({
                    'Status':               '',
                    'Usage':                f"{usage_val} Total",
                    'actor.properties.email': '',
                    'SUM of Total Usage':   usage_total,
                    '_row_type':            'usage_total'
                })
                status_total += usage_total

            # Status Total row
            rows.append({
                'Status':               f"{status_val} Total",
                'Usage':                '',
                'actor.properties.email': '',
                'SUM of Total Usage':   status_total,
                '_row_type':            'status_total'
            })
            grand_total += status_total

        # Grand Total row
        rows.append({
            'Status':               'Grand Total',
            'Usage':                '',
            'actor.properties.email': '',
            'SUM of Total Usage':   grand_total,
            '_row_type':            'grand_total'
        })

        return rows

    # ------------------------------------------------------------------
    # Other Insights (top free users)
    # ------------------------------------------------------------------
    def _build_other_insights(self):
        df = self.raw_df.copy()

        workspace_col = None
        for col in df.columns:
            for name in ['workspace', 'company', 'organization', 'account']:
                if name in str(col).lower():
                    workspace_col = col
                    break
            if workspace_col:
                break

        if not workspace_col:
            df['Workspace Name'] = 'Unknown'
            workspace_col = 'Workspace Name'

        def get_top_users(status_val, n=5):
            if 'Status' not in df.columns:
                return pd.DataFrame()
            subset = df[(df['Status'] == status_val) & (df.get('Status New', 'Leads') == 'Leads')].copy()
            if 'Status New' in df.columns:
                subset = df[(df['Status'] == status_val) & (df['Status New'] == 'Leads')].copy()
            if subset.empty:
                return pd.DataFrame()
            if 'Usage Count' in subset.columns:
                subset = subset.sort_values('Usage Count', ascending=False)
            return subset.head(n)

        top_new = get_top_users('NEW', 3)
        top_returning = get_top_users('RETURNING', 2)
        combined = pd.concat([top_new, top_returning])

        cols_to_keep = {
            workspace_col:           'Workspace Name',
            'Usage Count':           'Usage Count',
            'actor.properties.name': 'User Name',
            'clean_email':           'User Email',
        }
        final_cols = []
        for col, new_name in cols_to_keep.items():
            if col in combined.columns:
                final_cols.append(col)
            elif new_name == 'User Name':
                # Create blank column as placeholder; will be filled below
                combined[col] = ''
                final_cols.append(col)

        self.other_insights_df = combined[final_cols].rename(columns=cols_to_keep)

        # If User Name is missing or blank for any row, fall back to User Email
        # so the Other Insights table always shows something identifiable.
        if 'User Name' in self.other_insights_df.columns and 'User Email' in self.other_insights_df.columns:
            self.other_insights_df['User Name'] = self.other_insights_df.apply(
                lambda r: r['User Email'] if (pd.isna(r['User Name']) or str(r['User Name']).strip() == '') else r['User Name'],
                axis=1
            )
        print(f"Generated 'Other Insights' with {len(self.other_insights_df)} rows.")

    # ------------------------------------------------------------------
    # STEP 4: Save Report
    # ------------------------------------------------------------------
    def save_report(self):
        print(f"Saving report to {self.output_file}...")
        try:
            from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.dataframe import dataframe_to_rows

            # ---- Colour palette (matches screenshot) ----
            # Deep blue header (Status column)
            STATUS_HDR_FILL  = PatternFill("solid", fgColor="4472C4")   # Blue header
            STATUS_HDR_FONT  = Font(bold=True, color="FFFFFF")

            # Light lavender for pivot data rows
            DATA_FILL        = PatternFill("solid", fgColor="DCE6F1")   # Light blue-grey
            # Medium grey for usage subtotal rows
            USAGE_TOT_FILL   = PatternFill("solid", fgColor="B8CCE4")
            USAGE_TOT_FONT   = Font(bold=True)
            # Darker blue-grey for Status total rows
            STATUS_TOT_FILL  = PatternFill("solid", fgColor="8EA9C1")
            STATUS_TOT_FONT  = Font(bold=True)
            # Dark blue for Grand Total
            GRAND_TOT_FILL   = PatternFill("solid", fgColor="1F4E79")
            GRAND_TOT_FONT   = Font(bold=True, color="FFFFFF")

            # Generic header (for data sheets)
            HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
            HEADER_FONT = Font(bold=True)

            THIN = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            def auto_width(ws):
                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            def format_data_sheet(ws):
                """Apply border + header styling to a plain data sheet."""
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = THIN
                for cell in ws[1]:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                auto_width(ws)

            def write_pivot_sheet(wb, sheet_name, pivot_rows):
                """
                Writes a pivot table to a new sheet in the workbook.
                Columns: Status | Usage | actor.properties.email | SUM of Total Usage
                Styling matches the screenshot layout.
                """
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

                ws = wb.create_sheet(sheet_name)

                HEADERS = ['Status', 'Usage', 'actor.properties.email', 'SUM of Total Usage']

                # Write header row
                for col_idx, header in enumerate(HEADERS, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font   = STATUS_HDR_FONT
                    cell.fill   = STATUS_HDR_FILL
                    cell.border = THIN
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Write data rows
                for row_idx, row_data in enumerate(pivot_rows, start=2):
                    row_type = row_data.get('_row_type', 'data')

                    values = [
                        row_data.get('Status', ''),
                        row_data.get('Usage', ''),
                        row_data.get('actor.properties.email', ''),
                        row_data.get('SUM of Total Usage', ''),
                    ]

                    for col_idx, val in enumerate(values, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.border = THIN
                        cell.alignment = Alignment(vertical='center')

                        # Align numbers right
                        if col_idx == 4:
                            cell.alignment = Alignment(horizontal='right', vertical='center')

                        # Apply row-type specific styling
                        if row_type == 'grand_total':
                            cell.font = GRAND_TOT_FONT
                            cell.fill = GRAND_TOT_FILL
                        elif row_type == 'status_total':
                            cell.font = STATUS_TOT_FONT
                            cell.fill = STATUS_TOT_FILL
                        elif row_type == 'usage_total':
                            cell.font = USAGE_TOT_FONT
                            cell.fill = USAGE_TOT_FILL
                        else:
                            cell.fill = DATA_FILL

                auto_width(ws)
                # Fix: column D (SUM) should be narrower
                ws.column_dimensions['D'].width = 20
                print(f"  Pivot sheet '{sheet_name}' written with {len(pivot_rows)} rows.")

            # ---- Open writer ----
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:

                # 1. Main raw sheet
                self.main_sheet.to_excel(
                    writer,
                    sheet_name=f'Sheet {datetime.now().strftime("%d-%m-%Y")}',
                    index=False
                )

                # 2. Lead Sheet
                self.lead_sheet.to_excel(writer, sheet_name='Lead Sheet', index=False)

                # 3. Customer Sheet
                self.customer_sheet.to_excel(writer, sheet_name='Customer Sheet', index=False)

                # 4. Other Insights
                self.other_insights_df.to_excel(
                    writer, sheet_name='Other Insights', startrow=2, index=False
                )

                wb = writer.book

                # Apply formatting to data sheets
                for sname in [f'Sheet {datetime.now().strftime("%d-%m-%Y")}', 'Lead Sheet', 'Customer Sheet']:
                    if sname in wb.sheetnames:
                        format_data_sheet(wb[sname])

                # Other Insights custom headers
                if 'Other Insights' in wb.sheetnames:
                    oi_ws = wb['Other Insights']
                    oi_ws['A1'] = "Other Insights:"
                    oi_ws['A1'].font = Font(bold=True, size=12)
                    oi_ws['A2'] = "Some of the highlighted free users extensively using the product:"
                    oi_ws['A2'].font = Font(bold=True)
                    for cell in oi_ws[3]:
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.border = THIN
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    auto_width(oi_ws)

                # 5. Lead Pivot
                if self.lead_pivot_rows:
                    write_pivot_sheet(wb, 'Lead Pivot', self.lead_pivot_rows)
                else:
                    print("  No data for Lead Pivot — skipping sheet.")

                # 6. Customer Pivot
                if self.customer_pivot_rows:
                    write_pivot_sheet(wb, 'Customer Pivot', self.customer_pivot_rows)
                else:
                    print("  No data for Customer Pivot — skipping sheet.")

            print(f"Report saved successfully → {self.output_file}")
            return self.output_file

        except Exception as e:
            print(f"Error saving report: {e}")
            import traceback
            traceback.print_exc()
            return self.output_file

    # ------------------------------------------------------------------
    # STEP 5: Build Email HTML
    # ------------------------------------------------------------------
    def build_email_html(self, sheet_type='Lead', sheet_link='', prev_grand_total=0,
                          sender_name='', sender_title='', sender_linkedin='',
                          sender_website='', sender_email_addr='', sender_address='',
                          has_profile_image=False, has_award_image=False,
                          profile_image_bytes=None, award_image_bytes=None,
                          # Preview-mode params — inject editable toolbar when True
                          preview_mode=False, preview_id='', preview_password='',
                          preview_sheet_type='Lead',
                          preview_lead_recipients='',
                          preview_cust_recipients=''):
        """
        Build an inline-CSS HTML email body that mirrors the screenshot layout:
        - Greeting + intro line
        - Big grand-total number
        - % change from previous period
        - Google Sheet link
        - Pivot HTML table
        - Other Insights table
        - Signature block
        """
        if sheet_type == 'Lead':
            pivot_rows = self.lead_pivot_rows
        else:
            pivot_rows = self.customer_pivot_rows

        # ── Headline number: COUNT of unique active users in this sheet ──────
        # We count pivot rows of type 'data' (one per unique email), NOT the
        # SUM of usage values — that sum can be misleadingly large.
        user_count = sum(1 for r in pivot_rows if r.get('_row_type') == 'data')

        # Calculate % change vs previous period (only shown when prev is given)
        if prev_grand_total and prev_grand_total > 0:
            pct_change = ((user_count - prev_grand_total) / prev_grand_total) * 100
        else:
            pct_change = None  # hide change line entirely when no baseline

        pct_html = ''
        if pct_change is not None:
            pct_abs = abs(round(pct_change, 2))
            if pct_change < 0:
                change_arrow = '&#8600;'  # ↘
                change_color = '#e53935'
                change_text = f'Down {pct_abs}% from previous period'
            elif pct_change > 0:
                change_arrow = '&#8599;'  # ↗
                change_color = '#43a047'
                change_text = f'Up {pct_abs}% from previous period'
            else:
                change_arrow = '&#8594;'
                change_color = '#888888'
                change_text = 'No change from previous period'
            pct_html = f'''
            <p style="margin:0;font-size:14px;color:{change_color};">
              <span style="font-size:18px;">{change_arrow}</span>
              &nbsp;{change_text}
            </p>'''

        sheet_link_html = ''
        if sheet_link:
            sheet_link_html = f'''
            <p style="margin:10px 0 20px 0;font-size:14px;">
              <strong>For full sheet: </strong>
              <a href="{sheet_link}" style="color:#1a73e8;text-decoration:none;">Refer Here</a>
            </p>'''

        # Build pivot HTML table
        pivot_html = self._build_pivot_html_table(pivot_rows)

        # Build Other Insights table
        insights_html = self._build_insights_html_table()

        # Signature block
        profile_img_html = ''
        if has_profile_image:
            if preview_mode and profile_image_bytes:
                import base64
                b64_img = base64.b64encode(profile_image_bytes).decode('utf-8')
                src = f"data:image/jpeg;base64,{b64_img}"
            else:
                src = "cid:profile_image"
                
            profile_img_html = f'''
            <img src="{src}" alt="Profile Photo"
                 style="width:120px;height:120px;object-fit:cover;border-radius:4px;" />
            '''
        else:
            profile_img_html = '''
            <div style="width:120px;height:120px;background:#ccc;
                        border-radius:4px;display:flex;align-items:center;
                        justify-content:center;color:#666;font-size:12px;">Photo</div>
            '''

        award_img_html = ''
        if has_award_image:
            if preview_mode and award_image_bytes:
                import base64
                b64_aw = base64.b64encode(award_image_bytes).decode('utf-8')
                aw_src = f"data:image/png;base64,{b64_aw}"
            else:
                aw_src = "cid:award_image"
            award_img_html = f'<img src="{aw_src}" alt="Award" style="max-width:200px;margin-top:8px;" />'

        sig_linkedin = f'<a href="{sender_linkedin}" style="color:#0077b5;text-decoration:none;">{sender_name}</a>' if sender_linkedin else sender_name
        sig_website = f'<a href="{sender_website}" style="color:#333;text-decoration:none;">{sender_website}</a>' if sender_website else sender_website
        sig_email = f'<a href="mailto:{sender_email_addr}" style="color:#333;text-decoration:none;">{sender_email_addr}</a>' if sender_email_addr else sender_email_addr

        subject_label = f'{sheet_type} Report'

        html = f'''
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{subject_label}</title>
</head>
<body style="margin:0;padding:0;font-family:Arial,Helvetica,sans-serif;background:#f4f4f4;">

<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f4;padding:20px 0;">
  <tr>
    <td align="center">
      <table width="640" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:6px;
             border:1px solid #e0e0e0;overflow:hidden;">

        <!-- Header greeting -->
        <tr>
          <td style="padding:28px 32px 0 32px;">
            <p style="margin:0 0 6px 0;font-size:15px;color:#222;">Hello Everyone</p>
            <p style="margin:0 0 20px 0;font-size:14px;color:#444;">
              Please find the weekly <strong style="background:#fff3cd;padding:0 3px;">usage</strong>
              {sheet_type} report below:
            </p>
          </td>
        </tr>

        <!-- Active User Count -->
        <tr>
          <td style="padding:0 32px;">
            <p style="margin:0;font-size:72px;font-weight:900;color:#222;line-height:1.1;">{user_count}</p>
          </td>
        </tr>

        <!-- % Change (hidden when no previous period given) -->
        <tr>
          <td style="padding:4px 32px 8px 32px;">
            {pct_html}
          </td>
        </tr>

        <!-- Google Sheet link -->
        <tr>
          <td style="padding:0 32px;">
            {sheet_link_html}
          </td>
        </tr>

        <!-- Pivot Table -->
        <tr>
          <td style="padding:0 32px 20px 32px;">
            {pivot_html}
          </td>
        </tr>

        <!-- Other Insights -->
        <tr>
          <td style="padding:0 32px 24px 32px;">
            {insights_html}
          </td>
        </tr>

        <!-- Signature -->
        <tr>
          <td style="padding:16px 32px 28px 32px;border-top:1px solid #ebebeb;">
            <p style="margin:0 0 12px 0;font-size:14px;color:#333;">Best,</p>
            <table cellpadding="0" cellspacing="0" style="width:100%;">
              <tr>
                <td style="width:130px;vertical-align:top;">
                  {profile_img_html}
                </td>
                <td style="padding-left:16px;vertical-align:top;border-left:4px solid #2196f3;">
                  <p style="margin:0 0 2px 0;font-size:17px;font-weight:700;color:#2196f3;">{sender_name}</p>
                  <p style="margin:0 0 8px 0;font-size:13px;color:#555;">{sender_title}</p>
                  {'<p style="margin:2px 0;font-size:12px;color:#333;"><strong style="color:#0077b5;">Li</strong>&nbsp;&nbsp;' + sig_linkedin + '</p>' if sender_linkedin else ''}
                  {'<p style="margin:2px 0;font-size:12px;color:#333;"><strong>w:</strong>&nbsp;&nbsp;' + sig_website + '</p>' if sender_website else ''}
                  {'<p style="margin:2px 0;font-size:12px;color:#333;"><strong>e:</strong>&nbsp;&nbsp;' + sig_email + '</p>' if sender_email_addr else ''}
                  {'<p style="margin:2px 0;font-size:12px;color:#333;"><strong>a:</strong>&nbsp;&nbsp;' + sender_address + '</p>' if sender_address else ''}
                  {award_img_html}
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <!-- Confidentiality footer -->
        <tr>
          <td style="padding:10px 32px;background:#f9f9f9;border-top:1px solid #ebebeb;">
            <p style="margin:0;font-size:10px;color:#999;">
              IMPORTANT: The contents of this email and any attachments are confidential.
              It is strictly forbidden to share any part of this message without written consent.
            </p>
          </td>
        </tr>

      </table>
    </td>
  </tr>
</table>

</body>
</html>
'''
        if not preview_mode:
            return html


        # ── Preview Toolbar ───────────────────────────────────────────────
        # Injected at the top of the preview tab so the user can edit names
        # and send the email without returning to the main form.
        # Recipients are pre-populated from what was typed in the main form.
        default_recipients = (
            preview_lead_recipients if preview_sheet_type == 'Lead'
            else preview_cust_recipients
        )

        pwd_field = ''
        if preview_password:
            pwd_field = f'<input id="pvPassword" type="hidden" value="{preview_password}" />'
        else:
            pwd_field = '''
  <input id="pvPassword" type="password" placeholder="Gmail App Password"
    style="padding:8px 12px;border-radius:8px;border:none;font-size:13px;
           background:rgba(255,255,255,0.15);color:#fff;width:200px;
           outline:2px solid transparent;" />
            '''

        toolbar_html = f'''
<div id="previewToolbar" style="
    position:sticky;top:0;z-index:9999;
    background:linear-gradient(135deg,#2c3e6b,#4a6fa5);
    padding:14px 24px;display:flex;flex-wrap:wrap;
    align-items:center;gap:12px;
    box-shadow:0 3px 12px rgba(0,0,0,0.35);
    font-family:Arial,Helvetica,sans-serif;
">
  <span style="color:#fff;font-weight:700;font-size:14px;margin-right:4px;">
    ✉️ {preview_sheet_type} Preview
  </span>

  {pwd_field}

  <input id="pvRecipients" type="text"
    value="{default_recipients}"
    placeholder="recipient@example.com, ..."
    style="padding:8px 12px;border-radius:8px;border:none;font-size:13px;
           background:rgba(255,255,255,0.15);color:#fff;flex:1;min-width:200px;
           outline:2px solid transparent;" />

  <button onclick="sendFromPreview()" id="pvSendBtn"
    style="padding:9px 22px;border-radius:8px;border:none;
           background:#27ae60;color:#fff;font-weight:700;font-size:13px;
           cursor:pointer;transition:background 0.2s;white-space:nowrap;">
    📤 Send Now
  </button>

  <span id="pvStatus" style="color:#fff;font-size:13px;display:none;"></span>
</div>

<div style="font-family:Arial,sans-serif;font-size:12px;
     color:#555;text-align:center;padding:6px 0;
     background:#fffbea;border-bottom:1px solid #f0ad00;">
  ✏️ <strong>Edit mode active</strong> — click any <span
    style="background:#fffbea;border-bottom:2px dashed #f0ad00;padding:0 4px;"
  >highlighted cell</span> in the Other Insights table to fix names, then hit Send Now.
</div>

<script>
async function sendFromPreview() {{
  const password   = document.getElementById('pvPassword').value.trim().replace(/\\s/g,'');
  const recipients = document.getElementById('pvRecipients').value.trim();
  const btn        = document.getElementById('pvSendBtn');
  const status     = document.getElementById('pvStatus');

  if (!password)   {{ showPvStatus('❌ Enter your Gmail App Password.', '#e74c3c'); return; }}
  if (!recipients) {{ showPvStatus('❌ Enter at least one recipient.', '#e74c3c'); return; }}

  // Collect edited names and workspaces from the Other Insights table
  const nameOverrides = {{}};
  const workspaceOverrides = {{}};
  const rows = document.querySelectorAll('#insightsTable tbody tr');
  rows.forEach(tr => {{
    const email = tr.dataset.email;
    if (!email) return;
    
    const nameTd = tr.querySelector('td[data-field="User Name"]');
    if (nameTd) {{
      const editedName = nameTd.innerText.trim();
      if (editedName) nameOverrides[email] = editedName;
    }}
    
    const wsTd = tr.querySelector('td[data-field="Workspace Name"]');
    if (wsTd) {{
      const editedWs = wsTd.innerText.trim();
      if (editedWs) workspaceOverrides[email] = editedWs;
    }}
  }});

  btn.disabled = true;
  btn.textContent = '⏳ Sending…';
  showPvStatus('Sending email…', '#f39c12');

  try {{
    let targetOrigin = '';
    try {{ targetOrigin = window.parent.opener ? window.parent.opener.location.origin : window.parent.location.origin; }} catch(e) {{}}
    const targetUrl = targetOrigin && targetOrigin !== "null" ? targetOrigin + '/send_from_preview' : '/send_from_preview';

    const res = await fetch(targetUrl, {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify({{
        preview_id:     '{preview_id}',
        sheet_type:     '{preview_sheet_type}',
        sender_password: password,
        recipients:     recipients,
        name_overrides: nameOverrides,
        workspace_overrides: workspaceOverrides,
        sheet_link:     '',
      }})
    }});
    const data = await res.json();
    if (res.ok || res.status === 207) {{
      const sent = data.lead_sent || data.customer_sent;
      const errs = data.errors?.length ? ' | Errors: ' + data.errors.join(', ') : '';
      showPvStatus(sent ? '✅ Email sent!' + errs : '⚠️ Skipped' + errs,
                   sent ? '#27ae60' : '#e67e22');
    }} else {{
      showPvStatus('❌ ' + (data.error || 'Unknown error'), '#e74c3c');
    }}
  }} catch(err) {{
    showPvStatus('❌ Network error — ' + err.message, '#e74c3c');
  }} finally {{
    btn.disabled = false;
    btn.textContent = '📤 Send Now';
  }}
}}

function showPvStatus(msg, color) {{
  const el = document.getElementById('pvStatus');
  el.textContent = msg;
  el.style.color = color;
  el.style.display = 'inline';
}}
</script>
'''
        # Insert toolbar right after <body> tag
        return html.replace('<body', toolbar_html + '<body', 1)


    def _build_pivot_html_table(self, pivot_rows):
        """Render pivot_rows as an HTML table with inline styles."""
        if not pivot_rows:
            return '<p style="color:#888;font-size:13px;">No data available.</p>'

        header_style = (
            'background:#4472C4;color:#ffffff;font-weight:bold;'
            'padding:7px 10px;text-align:left;font-size:12px;border:1px solid #3a61a8;'
        )
        rows_html = f'''
        <table cellpadding="0" cellspacing="0" style="width:100%;border-collapse:collapse;font-size:12px;">
          <thead>
            <tr>
              <th style="{header_style}">status</th>
              <th style="{header_style}">usage</th>
              <th style="{header_style}">actor.properties.email</th>
              <th style="{header_style}text-align:right;">SUM of total usage</th>
            </tr>
          </thead>
          <tbody>
        '''

        for row in pivot_rows:
            rt = row.get('_row_type', 'data')
            if rt == 'grand_total':
                bg = '#1F4E79'; color = '#fff'; fw = 'bold'
            elif rt == 'status_total':
                bg = '#8EA9C1'; color = '#fff'; fw = 'bold'
            elif rt == 'usage_total':
                bg = '#B8CCE4'; color = '#222'; fw = 'bold'
            else:
                bg = '#DCE6F1'; color = '#222'; fw = 'normal'

            td = (
                f'background:{bg};color:{color};font-weight:{fw};'
                'padding:5px 8px;border:1px solid #ccc;'
            )
            email_val = row.get('actor.properties.email', '')
            # Make email a link if it looks like one
            if '@' in str(email_val):
                email_cell = f'<a href="mailto:{email_val}" style="color:#1a73e8;">{email_val}</a>'
            else:
                email_cell = email_val

            rows_html += f'''
            <tr>
              <td style="{td}">{row.get('Status', '')}</td>
              <td style="{td}">{row.get('Usage', '')}</td>
              <td style="{td}">{email_cell}</td>
              <td style="{td}text-align:right;">{row.get('SUM of Total Usage', '')}</td>
            </tr>
            '''

        rows_html += '</tbody></table>'
        return rows_html

    def _build_insights_html_table(self):
        """Render the Other Insights dataframe as an editable HTML table.
        
        In preview mode each 'User Name' and 'Workspace Name' cell is made
        contenteditable so the user can fix missing names before sending.
        A data-email attribute is attached to each row so the JS can map
        the edit back to the right user when building name_overrides.
        """
        if not hasattr(self, 'other_insights_df') or self.other_insights_df.empty:
            return ''

        df = self.other_insights_df.copy()

        header_style = (
            'background:#4472C4;color:#fff;font-weight:bold;'
            'padding:6px 10px;text-align:left;font-size:12px;border:1px solid #3a61a8;'
        )
        td_style      = 'padding:5px 8px;border:1px solid #ccc;font-size:12px;color:#222;'
        editable_style = (
            td_style +
            'background:#fffbea;outline:none;cursor:text;'
            'border-bottom:2px dashed #f0ad00;'
        )

        html = '''
        <p style="margin:20px 0 4px 0;font-size:14px;font-weight:bold;color:#222;">Other Insights:</p>
        <p style="margin:0 0 4px 0;font-size:11px;color:#888;">
          ✏️ <em>Click any highlighted cell to edit the name before sending.</em>
        </p>
        <p style="margin:0 0 8px 0;font-size:12px;color:#444;">
          Some of the highlighted free users extensively using the product:
        </p>
        <table cellpadding="0" cellspacing="0"
               style="width:100%;border-collapse:collapse;" id="insightsTable">
          <thead><tr>
        '''
        for col in df.columns:
            html += f'<th style="{header_style}">{col}</th>'
        html += '</tr></thead><tbody>'

        editable_cols = {'User Name', 'Workspace Name'}
        for row_idx, (_, row) in enumerate(df.iterrows()):
            email_val = row.get('User Email', '')
            html += f'<tr data-row="{row_idx}" data-email="{email_val}">'
            for col in df.columns:
                val = row[col]
                if col == 'User Email' and '@' in str(val):
                    cell = f'<a href="mailto:{val}" style="color:#1a73e8;">{val}</a>'
                    html += f'<td style="{td_style}">{cell}</td>'
                elif col in editable_cols:
                    # Editable cell — user can click and type
                    html += (
                        f'<td contenteditable="true" '
                        f'data-field="{col}" data-row="{row_idx}" '
                        f'style="{editable_style}">{str(val)}</td>'
                    )
                else:
                    html += f'<td style="{td_style}">{str(val)}</td>'
            html += '</tr>'

        html += '</tbody></table>'
        return html


# ----------------------------------------------------------------------
# Email Sender Utility
# ----------------------------------------------------------------------
def send_report_emails(
    generator,
    sender_email,
    sender_password,
    lead_recipients,
    customer_recipients,
    report_file_path=None,
    sheet_link='',
    prev_grand_total=0,
    sender_name='',
    sender_title='',
    sender_linkedin='',
    sender_website='',
    sender_email_addr='',
    sender_address='',
    profile_image_bytes=None,
    award_image_bytes=None
):
    """
    Sends Lead and Customer report emails via Gmail SMTP.
    Returns dict: { 'lead_sent': bool, 'customer_sent': bool, 'errors': [] }
    """
    results = {'lead_sent': False, 'customer_sent': False, 'errors': []}

    has_profile = profile_image_bytes is not None
    has_award = award_image_bytes is not None

    def _build_message(sheet_type, recipients):
        msg = MIMEMultipart('related')
        msg['From'] = f'{sender_name} <{sender_email}>' if sender_name else sender_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = (
            f'Weekly Usage Report — {sheet_type} Sheet '
            f'({datetime.now().strftime("%d %b %Y")})'
        )

        html_body = generator.build_email_html(
            sheet_type=sheet_type,
            sheet_link=sheet_link,
            prev_grand_total=prev_grand_total,
            sender_name=sender_name,
            sender_title=sender_title,
            sender_linkedin=sender_linkedin,
            sender_website=sender_website,
            sender_email_addr=sender_email_addr,
            sender_address=sender_address,
            has_profile_image=has_profile,
            has_award_image=has_award
        )

        # Wrap html in alternative part for plain-text fallback
        alt_part = MIMEMultipart('alternative')
        alt_part.attach(MIMEText('Please view this email in an HTML-capable client.', 'plain'))
        alt_part.attach(MIMEText(html_body, 'html'))
        msg.attach(alt_part)

        # Attach inline profile image
        if has_profile:
            img = MIMEImage(profile_image_bytes)
            img.add_header('Content-ID', '<profile_image>')
            img.add_header('Content-Disposition', 'inline', filename='profile.jpg')
            msg.attach(img)

        # Attach inline award image
        if has_award:
            img = MIMEImage(award_image_bytes)
            img.add_header('Content-ID', '<award_image>')
            img.add_header('Content-Disposition', 'inline', filename='award.png')
            msg.attach(img)

        # Attach Excel report
        if report_file_path and os.path.exists(report_file_path):
            with open(report_file_path, 'rb') as f:
                attachment = MIMEApplication(f.read(),
                    _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                attachment.add_header(
                    'Content-Disposition', 'attachment',
                    filename=os.path.basename(report_file_path)
                )
                msg.attach(attachment)

        return msg

    def _send(msg, recipients):
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipients, msg.as_string())

    # Send Lead email
    if lead_recipients:
        try:
            msg = _build_message('Lead', lead_recipients)
            _send(msg, lead_recipients)
            results['lead_sent'] = True
            print(f"Lead email sent to: {lead_recipients}")
        except Exception as e:
            results['errors'].append(f'Lead email failed: {e}')
            print(f"Lead email error: {e}")

    # Send Customer email
    if customer_recipients:
        try:
            msg = _build_message('Customer', customer_recipients)
            _send(msg, customer_recipients)
            results['customer_sent'] = True
            print(f"Customer email sent to: {customer_recipients}")
        except Exception as e:
            results['errors'].append(f'Customer email failed: {e}')
            print(f"Customer email error: {e}")

    return results


# ----------------------------------------------------------------------
# Entry Point
# ----------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    raw_data      = 'Untitled spreadsheet (1).xlsx'
    customer_list = 'customers.xlsx'

    if len(sys.argv) > 2:
        raw_data      = sys.argv[1]
        customer_list = sys.argv[2]

    if os.path.exists(raw_data) and os.path.exists(customer_list):
        print(f"Running with:\n  Raw Data:      {raw_data}\n  Customer List: {customer_list}")
        generator = ReportGenerator(raw_data, customer_list)
        generator.load_data()
        generator.process_data()
        generator.generate_sheets()
        generator.save_report()
    else:
        print(f"One or both files not found:\n  Raw: {raw_data}\n  Customer: {customer_list}")
